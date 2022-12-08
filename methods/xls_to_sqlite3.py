import sqlite3
import openpyxl
import os


def export_to_sqlite():
    # Экспорт данных из xlsx в sqlite

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'database.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS *Имя таблицы* (столбцы N штук, через ,)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('Путь к EXCEL файлу', data_only=True)
    sheet = file_to_read['Sheet1']

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 5 ( 5  включая) смотря сколько столбцов
        for col in range(1, 6):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
           

    # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO vacancy VALUES (?, ?, ?, ?, ? до [N-1]);", (data[0], data[1], data[2], data[3], data[4], "до data[N-1]"))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    # Очистка базы sqlite

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'database.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE")
    connect.commit()
    connect.close()



